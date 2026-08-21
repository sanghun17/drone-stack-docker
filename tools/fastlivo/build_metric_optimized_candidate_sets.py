#!/usr/bin/env python3
"""Build a 21-flight master table and metric-specific 5-vs-5 figure sets.

This is an exploratory *visualization-candidate* builder.  Recorded controller
conditions are retained as immutable provenance, but they are deliberately not
used by the optimizer.  For every endpoint, the optimizer finds five disjoint
``display-PURE -> display-Nominal`` pairs with:

* matched GT path length and flight duration;
* matched position-only opportunity to see the cloth AABB; and
* the largest favorable endpoint separation that satisfies arm-level balance.

Near-crash/manual-abort and otherwise incomplete flights stay in the master
table but are excluded from the default candidate pool.  The output labels are
therefore presentation roles, never replacements for recorded conditions.

All outcome metrics use the frozen stable-hover-to-landing window and the same
``full_livo_hybrid_imu_acc10_hover_r1`` replay for all 21 sessions.
"""

from __future__ import annotations

import argparse
import csv
from dataclasses import asdict, dataclass
import hashlib
import html
import json
import math
from pathlib import Path
import re
import sys
from typing import Any, Dict, Iterable, List, Mapping, Sequence, Tuple

import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt
from matplotlib.lines import Line2D
import numpy as np
import pandas as pd
from scipy.optimize import Bounds, LinearConstraint, milp


HERE = Path(__file__).resolve().parent
REPO = HERE.parents[1]
CAMPAIGN = HERE / "_campaign_20260805"
DEFAULT_OUTPUT = CAMPAIGN / "paper_metric_candidates_v1"
SPEC_PATH = HERE / "campaign_20260805_sessions.json"
RESULTS_PATH = (
    CAMPAIGN / "runs/full_livo_hybrid_imu_acc10_hover_r1/results.csv"
)
FUSION_PATH = (
    CAMPAIGN / "summary/full_livo_hybrid_imu_acc10_hover_r1/fusion_sessions.csv"
)
PREVIEW = (
    CAMPAIGN
    / "timeseries/production_primary/plots/paper_preview"
)
CLOTH_PATH = PREVIEW / "offline_common_cloth_s_sessions.csv"
SCENARIO_PATH = PREVIEW / "scenario_completion_audit.csv"
CACHE_INDEX_PATH = (
    CAMPAIGN / "timeseries/production_primary/cache/index.json"
)
INPUT_QC_PATH = CAMPAIGN / "input_summary/input_sessions.csv"

TOPIC_GT = "/vrpn_client_node/pure/pose"
TOPIC_EST = "/aft_mapped_to_optitrack"
MAX_ASSOCIATION_DIFF_S = 0.05
NEAR_CRASH_IDS = {
    "n0_20260805_021950",
    "n1_20260805_022212",
}

# Existing common offline cloth rasterizer and trajectory association helpers.
sys.path.insert(0, str(HERE))
from eval_fastlivo import (  # noqa: E402
    associate,
    geodesic_deg,
    q_to_R,
    read_traj,
    slerp,
)
from plot_flight_timeseries import _offline_cloth_view_fraction  # noqa: E402


CONDITION_COLORS = {
    "pure_wodz": "#2E86AB",
    "pure": "#8E44AD",
    "pure_mean": "#E67E22",
    "nominal": "#555555",
}
DISPLAY_PURE_COLOR = "#8E44AD"
DISPLAY_NOMINAL_COLOR = "#F2B134"


@dataclass(frozen=True)
class MetricSpec:
    column: str
    title: str
    unit: str
    higher_is_better: bool
    category: str
    note: str


METRICS: Tuple[MetricSpec, ...] = (
    MetricSpec(
        "ape_rmse_m", "Translation APE RMSE", "m", False, "estimation",
        "No spatial alignment; frozen per-session time offset.",
    ),
    MetricSpec(
        "ape_p90_m", "Translation APE p90", "m", False, "estimation",
        "90th percentile of associated translation APE.",
    ),
    MetricSpec(
        "ape_rmse_per_gt_path", "APE RMSE / GT path", "m/m", False,
        "estimation", "APE RMSE normalized by 10 Hz GT path length.",
    ),
    MetricSpec(
        "rpe1s_rmse_m", "1 s translation RPE RMSE", "m", False,
        "estimation", "World-frame one-second relative translation error.",
    ),
    MetricSpec(
        "rpe1s_p90_m", "1 s translation RPE p90", "m", False,
        "estimation", "90th percentile of one-second translation RPE.",
    ),
    MetricSpec(
        "orientation_rmse_deg", "Orientation error RMSE", "deg", False,
        "estimation", "SO(3) geodesic orientation error.",
    ),
    MetricSpec(
        "vio_nfeat_p10", "Visual support p10", "points", True, "tracking",
        "FAST-LIVO visual-map measurement support, not detector keypoints.",
    ),
    MetricSpec(
        "vio_nfeat_median", "Visual support median", "points", True,
        "tracking", "FAST-LIVO visual-map measurement support.",
    ),
    MetricSpec(
        "vio_nfeat_le_50_dwell_fraction", "Visual support <= 50 dwell",
        "fraction", False, "tracking",
        "Time-weighted fraction of VIO updates with support <= 50.",
    ),
    MetricSpec(
        "vio_effective_support_p10", "Improved-support proxy p10", "points",
        True, "tracking", "nfeat * clipped patch-improvement ratio.",
    ),
    MetricSpec(
        "vio_inlier_ratio_p10", "Patch-improvement ratio p10", "fraction",
        True, "tracking",
        "Fraction improved vs propagated-pose patch error; not RANSAC inliers.",
    ),
    MetricSpec(
        "vio_error_ratio_p90", "Photometric error ratio p90", "ratio", False,
        "tracking", "Final/propgated patch-error ratio; lower is better.",
    ),
    MetricSpec(
        "cloth_mean_fraction", "Mean cloth-view fraction", "fraction", False,
        "cloth", "Offline fixed-AABB/FOV geometric exposure.",
    ),
    MetricSpec(
        "cloth_p90_fraction", "Cloth-view fraction p90", "fraction", False,
        "cloth", "Offline fixed-AABB/FOV geometric exposure.",
    ),
    MetricSpec(
        "cloth_ge_0p50_dwell_fraction", "Cloth view >= 50% dwell", "fraction",
        False, "cloth", "Time-weighted high-exposure dwell fraction.",
    ),
    MetricSpec(
        "cloth_ge_0p70_dwell_fraction", "Cloth view >= 70% dwell", "fraction",
        False, "cloth", "Time-weighted very-high-exposure dwell fraction.",
    ),
    MetricSpec(
        "cloth_realization_over_opportunity", "Cloth view / position opportunity",
        "ratio", False, "cloth",
        "Mean actual-yaw exposure divided by max-yaw exposure opportunity.",
    ),
)


def sha256(path: Path, block_size: int = 8 << 20) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as stream:
        while True:
            block = stream.read(block_size)
            if not block:
                return digest.hexdigest()
            digest.update(block)
    return digest.hexdigest()


def atomic_json(path: Path, document: Mapping[str, Any]) -> None:
    temporary = path.with_suffix(path.suffix + ".tmp")
    temporary.write_text(
        json.dumps(document, indent=2, sort_keys=False, allow_nan=False),
        encoding="utf-8",
    )
    temporary.replace(path)


def finite_or_none(value: Any) -> Any:
    if isinstance(value, (np.floating, float)):
        return float(value) if math.isfinite(float(value)) else None
    if isinstance(value, (np.integer,)):
        return int(value)
    if isinstance(value, (np.bool_,)):
        return bool(value)
    return value


def json_records(frame: pd.DataFrame) -> List[Dict[str, Any]]:
    return [
        {str(key): finite_or_none(value) for key, value in row.items()}
        for row in frame.to_dict(orient="records")
    ]


def short_id(flight_id: str) -> str:
    return str(flight_id).split("_2026", 1)[0]


def quantile(values: np.ndarray, q: float) -> float:
    values = np.asarray(values, dtype=float)
    values = values[np.isfinite(values)]
    return float(np.quantile(values, q)) if len(values) else math.nan


def zoh_weights(times: np.ndarray, end: float | None = None) -> np.ndarray:
    """Duration represented by each sample, excluding an unobserved prefix."""
    times = np.asarray(times, dtype=float)
    if not len(times):
        return np.asarray([], dtype=float)
    order = np.argsort(times)
    if not np.array_equal(order, np.arange(len(times))):
        raise ValueError("zoh_weights requires sorted times")
    if end is None:
        tail = float(np.median(np.diff(times))) if len(times) > 1 else 0.1
        end = float(times[-1] + max(0.0, tail))
    edges = np.r_[times, max(float(end), float(times[-1]))]
    return np.maximum(0.0, np.diff(edges))


def weighted_fraction(mask: np.ndarray, weights: np.ndarray) -> float:
    mask = np.asarray(mask, dtype=bool)
    weights = np.asarray(weights, dtype=float)
    valid = np.isfinite(weights) & (weights >= 0.0)
    total = float(np.sum(weights[valid]))
    return float(np.sum(weights[valid] * mask[valid]) / total) if total else math.nan


def weighted_mean(values: np.ndarray, weights: np.ndarray) -> float:
    values = np.asarray(values, dtype=float)
    weights = np.asarray(weights, dtype=float)
    valid = np.isfinite(values) & np.isfinite(weights) & (weights >= 0.0)
    total = float(np.sum(weights[valid]))
    return float(np.sum(values[valid] * weights[valid]) / total) if total else math.nan


def count_events(times: np.ndarray, active: np.ndarray, merge_gap_s: float = 0.2) -> int:
    times = np.asarray(times, dtype=float)
    active = np.asarray(active, dtype=bool)
    indices = np.flatnonzero(active)
    if not len(indices):
        return 0
    count = 1
    last_t = float(times[indices[0]])
    for index in indices[1:]:
        current = float(times[index])
        if current - last_t > merge_gap_s:
            count += 1
        last_t = current
    return count


def translation_rpe_series(
    times: np.ndarray, estimate: np.ndarray, truth: np.ndarray, delta_s: float = 1.0
) -> np.ndarray:
    values: List[float] = []
    j = 1
    for i in range(len(times)):
        j = max(j, i + 1)
        while j < len(times) and times[j] - times[i] < delta_s:
            j += 1
        if j >= len(times):
            break
        values.append(float(np.linalg.norm(
            (estimate[j] - estimate[i]) - (truth[j] - truth[i])
        )))
    return np.asarray(values, dtype=float)


def trajectory_distributions(result_bag: Path, offset_s: float) -> Dict[str, float]:
    tg, xg, qg = read_traj(str(result_bag), TOPIC_GT)
    te, xe, qe = read_traj(str(result_bag), TOPIC_EST)
    pairs = associate(te + offset_s, tg, MAX_ASSOCIATION_DIFF_S)
    if len(pairs) < 10:
        raise RuntimeError(f"{result_bag}: only {len(pairs)} associations")
    pe = np.asarray([xe[i] for i, _, _, _ in pairs])
    pg = np.asarray([
        xg[k0] + u * (xg[k1] - xg[k0]) for _, k0, k1, u in pairs
    ])
    qgi = np.asarray([
        slerp(qg[k0], qg[k1], u) for _, k0, k1, u in pairs
    ])
    qe_assoc = np.asarray([qe[i] for i, _, _, _ in pairs])
    times = np.asarray([te[i] + offset_s for i, _, _, _ in pairs])
    elapsed = times - times[0]
    ape = np.linalg.norm(pe - pg, axis=1)
    rpe = translation_rpe_series(times, pe, pg, 1.0)
    ori = np.asarray([
        geodesic_deg(q_to_R(q_est), q_to_R(q_gt))
        for q_est, q_gt in zip(qe_assoc, qgi)
    ])
    tilt = np.asarray([
        math.degrees(math.acos(float(np.clip(
            np.dot(q_to_R(q_est)[:, 2], q_to_R(q_gt)[:, 2]), -1.0, 1.0
        ))))
        for q_est, q_gt in zip(qe_assoc, qgi)
    ])
    if len(elapsed) > 1 and elapsed[-1] > 0:
        ape_time_mean = float(np.trapz(ape, elapsed) / elapsed[-1])
        ape_slope = float(np.polyfit(elapsed, ape, 1)[0])
    else:
        ape_time_mean = float(np.mean(ape))
        ape_slope = math.nan
    return {
        "ape_association_count": int(len(ape)),
        "ape_median_m": float(np.median(ape)),
        "ape_p90_m": quantile(ape, 0.90),
        "ape_p95_m": quantile(ape, 0.95),
        "ape_time_mean_m": ape_time_mean,
        "ape_slope_mps": ape_slope,
        "ape_end_minus_start_m": float(ape[-1] - ape[0]),
        "rpe1s_count": int(len(rpe)),
        "rpe1s_median_m": float(np.median(rpe)),
        "rpe1s_p90_m": quantile(rpe, 0.90),
        "rpe1s_p95_m": quantile(rpe, 0.95),
        "rpe1s_max_m": float(np.max(rpe)),
        "orientation_median_deg": float(np.median(ori)),
        "orientation_p95_deg": quantile(ori, 0.95),
        "tilt_axis_median_deg": float(np.median(tilt)),
        "tilt_axis_p95_deg": quantile(tilt, 0.95),
    }


def stage_feature_metrics(
    frame: pd.DataFrame, stage: str, duration_s: float
) -> Dict[str, float]:
    rows = frame.loc[frame["stage"].astype(str) == stage].copy()
    rows = rows.loc[(rows["t"] >= 0.0) & (rows["t"] <= duration_s)]
    rows = rows.sort_values("t").drop_duplicates("t", keep="last")
    prefix = stage.lower()
    if rows.empty:
        return {f"{prefix}_rows": 0}
    times = rows["t"].to_numpy(dtype=float)
    weights = zoh_weights(times, duration_s)
    nfeat = rows["nfeat"].to_numpy(dtype=float)
    span = max(0.0, float(times[-1] - times[0]))
    result: Dict[str, float] = {
        f"{prefix}_rows": int(len(rows)),
        f"{prefix}_update_hz": float(len(rows) / max(span, 1e-9)),
        f"{prefix}_time_span_s": span,
        f"{prefix}_span_fraction": float(span / max(duration_s, 1e-9)),
        f"{prefix}_nfeat_min": float(np.min(nfeat)),
        f"{prefix}_nfeat_p05": quantile(nfeat, 0.05),
        f"{prefix}_nfeat_p10": quantile(nfeat, 0.10),
        f"{prefix}_nfeat_p25": quantile(nfeat, 0.25),
        f"{prefix}_nfeat_median": float(np.median(nfeat)),
        f"{prefix}_nfeat_mean": float(np.mean(nfeat)),
        f"{prefix}_nfeat_p75": quantile(nfeat, 0.75),
        f"{prefix}_nfeat_p90": quantile(nfeat, 0.90),
        f"{prefix}_nfeat_max": float(np.max(nfeat)),
        f"{prefix}_nfeat_zero_dwell_fraction": weighted_fraction(nfeat <= 0, weights),
    }
    for threshold in (20, 30, 50, 75, 100):
        result[f"{prefix}_nfeat_le_{threshold}_dwell_fraction"] = weighted_fraction(
            nfeat <= threshold, weights
        )

    # The following quantities are populated on both stage rows by the replay,
    # but only VIO-stage values are interpreted as visual-update diagnostics.
    if stage == "VIO":
        inlier = rows["vio_inlier_ratio"].to_numpy(dtype=float)
        error = rows["vio_error_ratio"].to_numpy(dtype=float)
        effective = nfeat * np.clip(inlier, 0.0, 1.0)
        supported = nfeat > 0
        result.update({
            "vio_effective_support_p10": quantile(effective, 0.10),
            "vio_effective_support_median": float(np.median(effective)),
            "vio_effective_support_mean": float(np.mean(effective)),
            "vio_effective_support_le_50_dwell_fraction": weighted_fraction(
                effective <= 50.0, weights
            ),
            "vio_supported_update_fraction": weighted_fraction(supported, weights),
        })
        for name, values, qs in (
            ("vio_inlier_ratio", inlier[supported], (0.10, 0.50, 0.90)),
            ("vio_error_ratio", error[supported], (0.10, 0.50, 0.90)),
            ("vio_trans_info_ratio",
             rows.loc[supported, "vio_trans_info_ratio"].to_numpy(float),
             (0.10, 0.50)),
            ("vio_rot_info_ratio",
             rows.loc[supported, "vio_rot_info_ratio"].to_numpy(float),
             (0.10, 0.50)),
            ("vio_info_min_per_measurement",
             rows.loc[supported, "vio_info_min_per_measurement"].to_numpy(float),
             (0.10, 0.50)),
        ):
            values = values[np.isfinite(values)]
            for q in qs:
                suffix = "median" if q == 0.50 else f"p{int(round(100*q)):02d}"
                result[f"{name}_{suffix}"] = quantile(values, q)
            if len(values):
                result[f"{name}_mean"] = float(np.mean(values))
    elif stage == "LIO":
        for name in (
            "lio_trans_info_ratio",
            "lio_rot_info_ratio",
            "lio_info_min_per_feature",
        ):
            values = rows[name].to_numpy(dtype=float)
            values = values[np.isfinite(values)]
            result[f"{name}_p10"] = quantile(values, 0.10)
            result[f"{name}_median"] = quantile(values, 0.50)
    return result


def interp_columns(times: np.ndarray, values: np.ndarray, query: np.ndarray) -> np.ndarray:
    return np.column_stack([
        np.interp(query, times, values[:, axis]) for axis in range(values.shape[1])
    ])


def opportunity_curve(
    positions: np.ndarray, yaw_step_deg: float = 5.0, position_batch: int = 16
) -> np.ndarray:
    """Maximum cloth FOV-fill over a deterministic yaw grid at each position."""
    positions = np.asarray(positions, dtype=float)
    yaw_grid = np.deg2rad(np.arange(-180.0, 180.0, yaw_step_deg, dtype=float))
    result = np.empty(len(positions), dtype=float)
    for start in range(0, len(positions), position_batch):
        stop = min(len(positions), start + position_batch)
        block = positions[start:stop]
        count = len(block)
        poses = np.empty((count * len(yaw_grid), 4), dtype=float)
        poses[:, :3] = np.repeat(block, len(yaw_grid), axis=0)
        poses[:, 3] = np.tile(yaw_grid, count)
        fraction = _offline_cloth_view_fraction(poses).reshape(count, -1)
        result[start:stop] = np.max(fraction, axis=1)
    return result


def cloth_metrics(
    npz_path: Path,
    start_s: float,
    duration_s: float,
    opportunity_cache: Path,
    yaw_step_deg: float,
) -> Dict[str, float]:
    with np.load(npz_path) as data:
        relative = data["time_from_bag_start_s"].astype(float)
        poses = data["pose_xyz_yaw"].astype(float)
        fraction = data["cloth_view_fraction"].astype(float)
    end_s = start_s + duration_s
    keep = (relative >= start_s) & (relative <= end_s)
    if np.count_nonzero(keep) < 10:
        raise RuntimeError(f"{npz_path}: no stable-hover window")
    t = relative[keep]
    xyz = poses[keep, :3]
    f = fraction[keep]
    elapsed = t - t[0]
    weights = zoh_weights(elapsed, float(elapsed[-1]))
    result: Dict[str, float] = {
        "cloth_samples": int(len(f)),
        "cloth_mean_fraction": weighted_mean(f, weights),
        "cloth_median_fraction": float(np.median(f)),
        "cloth_p90_fraction": quantile(f, 0.90),
        "cloth_p95_fraction": quantile(f, 0.95),
        "cloth_peak_fraction": float(np.max(f)),
        "cloth_auc_fraction_s": float(np.sum(f * weights)),
        "cloth_active_dwell_fraction": weighted_fraction(f > 0.0, weights),
    }
    for threshold, tag in ((0.25, "0p25"), (0.50, "0p50"),
                           (0.70, "0p70"), (0.75, "0p75")):
        active = f >= threshold
        result[f"cloth_ge_{tag}_dwell_fraction"] = weighted_fraction(active, weights)
        result[f"cloth_ge_{tag}_event_count"] = count_events(elapsed, active)

    # Path-weighted exposure uses a fixed 10 Hz interpolation to avoid mocap
    # jitter dominating cumulative distance.
    query = np.arange(float(t[0]), float(t[-1]) + 1e-9, 0.1)
    xyz10 = interp_columns(t, xyz, query)
    f10 = np.interp(query, t, f)
    segment = np.linalg.norm(np.diff(xyz10, axis=0), axis=1)
    total_path = float(np.sum(segment))
    mid_f = 0.5 * (f10[:-1] + f10[1:])
    for threshold, tag in ((0.25, "0p25"), (0.50, "0p50"), (0.70, "0p70")):
        result[f"cloth_ge_{tag}_path_fraction"] = (
            float(np.sum(segment[mid_f >= threshold]) / total_path)
            if total_path > 0 else math.nan
        )

    opportunity_cache.parent.mkdir(parents=True, exist_ok=True)
    cache_ok = False
    if opportunity_cache.is_file():
        try:
            with np.load(opportunity_cache) as cached:
                cache_ok = (
                    abs(float(cached["start_s"]) - start_s) < 1e-9
                    and abs(float(cached["duration_s"]) - duration_s) < 1e-9
                    and abs(float(cached["yaw_step_deg"]) - yaw_step_deg) < 1e-9
                    and np.array_equal(cached["query_s"], query)
                )
                if cache_ok:
                    opportunity = cached["opportunity_fraction"].copy()
        except Exception:
            cache_ok = False
    if not cache_ok:
        opportunity = opportunity_curve(xyz10, yaw_step_deg=yaw_step_deg)
        np.savez_compressed(
            opportunity_cache,
            query_s=query,
            opportunity_fraction=opportunity,
            start_s=np.asarray(start_s),
            duration_s=np.asarray(duration_s),
            yaw_step_deg=np.asarray(yaw_step_deg),
        )
    opportunity_weights = zoh_weights(query - query[0], float(query[-1] - query[0]))
    opportunity_mean = weighted_mean(opportunity, opportunity_weights)
    result.update({
        "cloth_position_opportunity_mean_fraction": opportunity_mean,
        "cloth_position_opportunity_p90_fraction": quantile(opportunity, 0.90),
        "cloth_position_opportunity_peak_fraction": float(np.max(opportunity)),
        "cloth_realization_over_opportunity": (
            result["cloth_mean_fraction"] / opportunity_mean
            if opportunity_mean > 1e-12 else math.nan
        ),
        "cloth_opportunity_yaw_grid_step_deg": yaw_step_deg,
        "cloth_opportunity_position_hz": 10.0,
    })
    return result


def build_master(output: Path, yaw_step_deg: float) -> Tuple[pd.DataFrame, Dict[str, Any]]:
    catalog_doc = json.loads(SPEC_PATH.read_text(encoding="utf-8"))
    catalog = {str(row["id"]): row for row in catalog_doc["sessions"]}
    recordings_root = Path(catalog_doc["recordings_root"])
    results = pd.read_csv(RESULTS_PATH).set_index("flight_id", drop=False)
    fusion = pd.read_csv(FUSION_PATH).set_index("flight_id", drop=False)
    cloth = pd.read_csv(CLOTH_PATH).set_index("flight_id", drop=False)
    scenario = pd.read_csv(SCENARIO_PATH).set_index("flight_id", drop=False)
    input_qc = pd.read_csv(INPUT_QC_PATH).set_index("flight_id", drop=False)
    cache_doc = json.loads(CACHE_INDEX_PATH.read_text(encoding="utf-8"))
    cache_rows = {
        str(row["flight_id"]): row for row in cache_doc["sessions"]
    }
    rows: List[Dict[str, Any]] = []
    opportunity_dir = output / "opportunity_cache"
    ordered_ids = [str(row["id"]) for row in catalog_doc["sessions"]]
    for index, flight_id in enumerate(ordered_ids, start=1):
        print(f"[master {index:02d}/{len(ordered_ids)}] {flight_id}", flush=True)
        src = catalog[flight_id]
        r = results.loc[flight_id]
        fsum = fusion.loc[flight_id]
        c = cloth.loc[flight_id]
        s = scenario.loc[flight_id]
        q = input_qc.loc[flight_id]
        cache = cache_rows[flight_id]
        source = Path(str(src["source"]))
        if not source.is_absolute():
            source = recordings_root / source
        duration = float(r["duration_s"])
        endpoint_complete = bool(
            float(s["start_distance_m"]) <= 0.5
            and bool(s["terminal_nav_observed"])
            and bool(s["done_observed"])
            and bool(s["landing_observed_after_hover"])
            and bool(s["configured_terminal_gt_within_0p2m"])
        )
        result: Dict[str, Any] = {
            "flight_id": flight_id,
            "session_id": str(r["session_id"]),
            "recorded_condition": str(src["condition"]),
            "split": str(src["split"]),
            "source_bag": str(source),
            "source_present": source.is_file(),
            "canonical_bag": str(r["canonical_bag"]),
            "input_bag": str(r["input_bag"]),
            "result_bag": str(r["result_bag"]),
            "fusion_csv": str(fsum["fusion_csv"]),
            "window_start_from_bag_s": float(r["start_s"]),
            "window_duration_s": duration,
            "window_method": str(cache["windows"]["hover"]["method"]),
            "time_offset_s": float(r["time_offset_s"]),
            "estimator_overlay_config": str(r["overlay"]),
            "estimator_rate": float(r["rate"]),
            "near_crash_manual_label": flight_id in NEAR_CRASH_IDS,
            "endpoint_complete": endpoint_complete,
            "censored_or_incomplete": not endpoint_complete,
            "completion_reason": str(s["reason"]),
            "completion_start_distance_m": float(s["start_distance_m"]),
            "completion_terminal_nav_observed": bool(s["terminal_nav_observed"]),
            "completion_done_observed": bool(s["done_observed"]),
            "completion_landing_observed": bool(s["landing_observed_after_hover"]),
            "completion_terminal_within_0p2m": bool(
                s["configured_terminal_gt_within_0p2m"]
            ),
            "completion_closest_terminal_distance_m": float(
                s["closest_configured_terminal_xyz_distance_m"]
            ),
            "completion_scenario_duration_s": float(s["scenario_duration_s"]),
            "completion_landing_minus_done_s": float(s["landing_minus_done_s"]),
            "gt_path_m": float(r["gt_path_m"]),
            "gt_associated_path_m": float(r["gt_associated_path_m"]),
            "gt_duration_s": float(r["gt_duration_s"]),
            "gt_mean_speed_mps": float(r["gt_path_m"] / max(duration, 1e-9)),
            "est_path_m": float(r["est_path_m"]),
            "motion_ratio": float(r["motion_ratio"]),
            "ape_rmse_m": float(r["rmse_m"]),
            "ape_mean_m": float(r["mean_ape_m"]),
            "ape_max_m": float(r["max_ape_m"]),
            "ape_final_m": float(r["final_ape_m"]),
            "ape_rmse_per_gt_path": float(r["rmse_per_gt_path"]),
            "rpe1s_rmse_m": float(r["rpe_1s_rmse_m"]),
            "orientation_rmse_deg": float(r["orientation_rmse_deg"]),
            "orientation_mean_deg": float(r["orientation_mean_deg"]),
            "orientation_p90_deg": float(r["orientation_p90_deg"]),
            "orientation_max_deg": float(r["orientation_max_deg"]),
            "tilt_axis_rmse_deg": float(r["tilt_axis_rmse_deg"]),
            "tilt_axis_p90_deg": float(r["tilt_axis_p90_deg"]),
            "trend_1s_weighted_cosine": float(r["trend_1s_weighted_cosine"]),
            "trend_reverse_distance_fraction": float(
                r["trend_reverse_distance_fraction"]
            ),
            "trend_stall_fraction": float(r["trend_stall_fraction"]),
            "trend_progress_correlation": float(r["trend_progress_correlation"]),
            "net_displacement_cosine": float(r["net_displacement_cosine"]),
            "net_displacement_ratio": float(r["net_displacement_ratio"]),
            "estimator_associations": int(r["associations"]),
            "estimator_coverage": float(r["coverage"]),
            "estimator_output_count": int(r["output_count"]),
            "estimator_output_ratio": float(r["output_ratio"]),
            "estimator_max_output_gap_s": float(r["max_output_gap_s"]),
            "estimator_valid": bool(r["valid"]),
            "estimator_integrity": bool(r["integrity"]),
            "estimator_trend_consistent": bool(r["trend_consistent"]),
            "estimator_catastrophic": bool(r["catastrophic"]),
        }
        result.update(trajectory_distributions(
            Path(str(r["result_bag"])), float(r["time_offset_s"])
        ))
        fusion_frame = pd.read_csv(str(fsum["fusion_csv"]))
        result.update(stage_feature_metrics(fusion_frame, "LIO", duration))
        result.update(stage_feature_metrics(fusion_frame, "VIO", duration))
        result.update(cloth_metrics(
            Path(str(c["offline_series_npz"])),
            float(r["start_s"]), duration,
            opportunity_dir / f"{flight_id}.npz", yaw_step_deg,
        ))
        for key, value in q.items():
            if key in {"flight_id", "condition", "split"}:
                continue
            result[f"inputqc_fullbag_{key}"] = value
        feature_ok = bool(
            int(result.get("vio_rows", 0)) > 0
            and int(result.get("lio_rows", 0)) > 0
            and float(result.get("vio_span_fraction", 0.0)) >= 0.95
            and math.isfinite(float(result.get("vio_nfeat_p10", math.nan)))
        )
        cloth_ok = bool(
            int(result.get("cloth_samples", 0)) > 10
            and 0.0 <= float(result.get("cloth_mean_fraction", math.nan)) <= 1.0
        )
        estimator_ok = bool(
            result["estimator_valid"]
            and result["estimator_integrity"]
            and result["estimator_trend_consistent"]
            and result["estimator_coverage"] >= 0.95
        )
        result.update({
            "eligible_estimator": estimator_ok,
            "eligible_feature": feature_ok,
            "eligible_cloth": cloth_ok,
            "eligible_primary_selection": bool(
                estimator_ok and feature_ok and cloth_ok and endpoint_complete
            ),
        })
        rows.append(result)

    master = pd.DataFrame(rows)
    # Stable column families make the wide CSV easier to navigate.
    family_order = (
        "flight_id", "session_id", "recorded_condition", "split",
        "source_bag", "source_present", "canonical_bag", "input_bag",
        "result_bag", "fusion_csv", "window_", "near_crash_", "endpoint_",
        "censored_", "completion_", "eligible_", "gt_", "est_path_",
        "motion_", "ape_", "rpe1s_", "orientation_", "tilt_", "trend_",
        "net_", "estimator_", "lio_", "vio_", "cloth_", "inputqc_",
        "time_offset_",
    )
    columns: List[str] = []
    for prefix in family_order:
        for column in master.columns:
            if column not in columns and (
                column == prefix or column.startswith(prefix)
            ):
                columns.append(column)
    columns.extend(column for column in master.columns if column not in columns)
    master = master[columns]
    metadata = {
        "session_count": len(master),
        "eligible_primary_count": int(master["eligible_primary_selection"].sum()),
        "incomplete_flight_ids": master.loc[
            ~master["endpoint_complete"], "flight_id"
        ].tolist(),
        "near_crash_ids": sorted(NEAR_CRASH_IDS),
    }
    return master, metadata


def optimize_pairs(
    master: pd.DataFrame,
    metric: MetricSpec,
    pair_count: int = 5,
) -> Dict[str, Any]:
    pool = master.loc[master["eligible_primary_selection"]].copy()
    pool = pool.loc[np.isfinite(pool[metric.column].astype(float))].reset_index(drop=True)
    if len(pool) < 2 * pair_count:
        raise RuntimeError(f"{metric.column}: only {len(pool)} eligible rows")
    metric_values = pool[metric.column].to_numpy(dtype=float)
    metric_sd = float(np.std(metric_values, ddof=0))
    if metric_sd <= 1e-12:
        raise RuntimeError(f"{metric.column}: zero variance")
    covariates = np.column_stack([
        np.log(pool["gt_path_m"].to_numpy(dtype=float)),
        np.log(pool["window_duration_s"].to_numpy(dtype=float)),
        pool["cloth_position_opportunity_mean_fraction"].to_numpy(dtype=float),
    ])
    covariate_names = ("log_gt_path", "log_duration", "position_opportunity")
    cov_sd = np.std(covariates, axis=0, ddof=0)
    cov_sd = np.where(cov_sd > 1e-12, cov_sd, 1.0)
    edges: List[Dict[str, Any]] = []
    for a in range(len(pool)):
        for b in range(a + 1, len(pool)):
            va, vb = metric_values[a], metric_values[b]
            if abs(va - vb) <= 1e-12:
                continue
            if metric.higher_is_better:
                good, bad = (a, b) if va > vb else (b, a)
            else:
                good, bad = (a, b) if va < vb else (b, a)
            path_ratio = max(
                float(pool.loc[good, "gt_path_m"] / pool.loc[bad, "gt_path_m"]),
                float(pool.loc[bad, "gt_path_m"] / pool.loc[good, "gt_path_m"]),
            )
            duration_ratio = max(
                float(pool.loc[good, "window_duration_s"] /
                      pool.loc[bad, "window_duration_s"]),
                float(pool.loc[bad, "window_duration_s"] /
                      pool.loc[good, "window_duration_s"]),
            )
            opportunity_gap = abs(float(covariates[good, 2] - covariates[bad, 2]))
            if path_ratio > 1.20 or duration_ratio > 1.25 or opportunity_gap > 0.10:
                continue
            favorable_gap = (
                metric_values[good] - metric_values[bad]
                if metric.higher_is_better
                else metric_values[bad] - metric_values[good]
            )
            cov_diff = covariates[good] - covariates[bad]
            pair_cost = float(np.sum(np.abs(cov_diff) / cov_sd))
            edges.append({
                "good": good,
                "bad": bad,
                "benefit_z": float(favorable_gap / metric_sd),
                "cov_diff": cov_diff,
                "pair_cost": pair_cost,
                "path_ratio": path_ratio,
                "duration_ratio": duration_ratio,
                "opportunity_gap": opportunity_gap,
            })
    if not edges:
        raise RuntimeError(f"{metric.column}: no caliper-compatible pairs")

    count = len(edges)
    constraints: List[LinearConstraint] = []
    constraints.append(LinearConstraint(np.ones((1, count)), [pair_count], [pair_count]))
    incidence = np.zeros((len(pool), count), dtype=float)
    for edge_index, edge in enumerate(edges):
        incidence[edge["good"], edge_index] = 1.0
        incidence[edge["bad"], edge_index] = 1.0
    constraints.append(LinearConstraint(
        incidence, np.zeros(len(pool)), np.ones(len(pool))
    ))

    selected_indices: np.ndarray | None = None
    used_balance_delta = math.nan
    result_obj = None
    for balance_delta in (0.10, 0.15, 0.20, 0.30, 0.50, 1.0):
        balance = np.column_stack([edge["cov_diff"] for edge in edges])
        bound = pair_count * balance_delta * cov_sd
        balance_constraint = LinearConstraint(balance, -bound, bound)
        # Deterministic tie breaking: covariate distance, then edge order.
        objective = np.asarray([
            -edge["benefit_z"]
            + 1e-4 * edge["pair_cost"]
            + 1e-9 * edge_index
            for edge_index, edge in enumerate(edges)
        ])
        solved = milp(
            objective,
            integrality=np.ones(count, dtype=int),
            bounds=Bounds(np.zeros(count), np.ones(count)),
            constraints=[*constraints, balance_constraint],
            options={"time_limit": 60.0, "mip_rel_gap": 0.0},
        )
        if solved.success and solved.x is not None:
            choice = np.flatnonzero(solved.x > 0.5)
            if len(choice) == pair_count:
                selected_indices = choice
                used_balance_delta = balance_delta
                result_obj = solved
                break
    if selected_indices is None:
        raise RuntimeError(f"{metric.column}: MILP infeasible after balance relaxation")

    pair_rows: List[Dict[str, Any]] = []
    good_values: List[float] = []
    bad_values: List[float] = []
    good_cov: List[np.ndarray] = []
    bad_cov: List[np.ndarray] = []
    selected_edges = [edges[index] for index in selected_indices]
    selected_edges.sort(key=lambda edge: str(pool.loc[edge["good"], "flight_id"]))
    for pair_index, edge in enumerate(selected_edges, start=1):
        good = pool.loc[edge["good"]]
        bad = pool.loc[edge["bad"]]
        good_value = float(good[metric.column])
        bad_value = float(bad[metric.column])
        good_values.append(good_value)
        bad_values.append(bad_value)
        good_cov.append(covariates[edge["good"]])
        bad_cov.append(covariates[edge["bad"]])
        pair_rows.append({
            "metric": metric.column,
            "pair_index": pair_index,
            "display_pure_flight_id": str(good["flight_id"]),
            "display_pure_short_id": short_id(str(good["flight_id"])),
            "display_pure_recorded_condition": str(good["recorded_condition"]),
            "display_pure_value": good_value,
            "display_nominal_flight_id": str(bad["flight_id"]),
            "display_nominal_short_id": short_id(str(bad["flight_id"])),
            "display_nominal_recorded_condition": str(bad["recorded_condition"]),
            "display_nominal_value": bad_value,
            "favorable_gap": (
                good_value - bad_value if metric.higher_is_better
                else bad_value - good_value
            ),
            "gt_path_ratio_max_over_min": edge["path_ratio"],
            "duration_ratio_max_over_min": edge["duration_ratio"],
            "position_opportunity_abs_gap": edge["opportunity_gap"],
        })
    good_values_array = np.asarray(good_values)
    bad_values_array = np.asarray(bad_values)
    good_cov_array = np.asarray(good_cov)
    bad_cov_array = np.asarray(bad_cov)
    favorable_mean_gap = (
        float(np.mean(good_values_array) - np.mean(bad_values_array))
        if metric.higher_is_better
        else float(np.mean(bad_values_array) - np.mean(good_values_array))
    )
    cov_smd = np.abs(
        (np.mean(good_cov_array, axis=0) - np.mean(bad_cov_array, axis=0)) / cov_sd
    )
    return {
        "metric": asdict(metric),
        "eligible_pool_size": int(len(pool)),
        "pair_count": pair_count,
        "pair_calipers": {
            "gt_path_ratio_max": 1.20,
            "duration_ratio_max": 1.25,
            "position_opportunity_abs_gap_max": 0.10,
        },
        "arm_balance_delta_smd": used_balance_delta,
        "display_pure_mean": float(np.mean(good_values_array)),
        "display_nominal_mean": float(np.mean(bad_values_array)),
        "favorable_mean_gap": favorable_mean_gap,
        "favorable_separation_pool_sd": float(favorable_mean_gap / metric_sd),
        "covariate_smd": {
            name: float(value) for name, value in zip(covariate_names, cov_smd)
        },
        "objective_value": float(result_obj.fun) if result_obj is not None else None,
        "pairs": pair_rows,
    }


def metric_formatter(metric: MetricSpec, value: float) -> str:
    if metric.unit == "fraction":
        return f"{100.0 * value:.1f}%"
    if metric.unit == "deg":
        return f"{value:.1f} deg"
    if metric.unit == "points":
        return f"{value:.1f}"
    if metric.unit in {"ratio", "m/m"}:
        return f"{value:.3f}"
    return f"{value:.3f} {metric.unit}".strip()


def spread_label_positions(
    values: Sequence[float], lower: float, upper: float, min_gap: float
) -> np.ndarray:
    """Return order-preserving label y positions with a guaranteed separation."""
    values_array = np.asarray(values, dtype=float)
    order = np.argsort(values_array)
    placed = values_array[order].copy()
    for index in range(1, len(placed)):
        placed[index] = max(placed[index], placed[index - 1] + min_gap)
    if len(placed) and placed[-1] > upper:
        placed -= placed[-1] - upper
    for index in range(len(placed) - 2, -1, -1):
        placed[index] = min(placed[index], placed[index + 1] - min_gap)
    if len(placed) and placed[0] < lower:
        placed += lower - placed[0]
    result = np.empty_like(placed)
    result[order] = placed
    return result


def plot_metric(
    master: pd.DataFrame,
    metric: MetricSpec,
    selection: Mapping[str, Any],
    plots_dir: Path,
) -> None:
    plots_dir.mkdir(parents=True, exist_ok=True)
    selected_good = {
        row["display_pure_flight_id"] for row in selection["pairs"]
    }
    selected_bad = {
        row["display_nominal_flight_id"] for row in selection["pairs"]
    }
    finite = master.loc[np.isfinite(master[metric.column].astype(float))].copy()
    finite = finite.sort_values(metric.column, ascending=not metric.higher_is_better)
    fig, (ax_all, ax_pairs) = plt.subplots(
        1, 2, figsize=(14.2, 6.4), gridspec_kw={"width_ratios": [1.25, 1.0]}
    )
    for x, (_, row) in enumerate(finite.iterrows()):
        fid = str(row["flight_id"])
        eligible = bool(row["eligible_primary_selection"])
        color = CONDITION_COLORS.get(str(row["recorded_condition"]), "#777777")
        marker = "o" if eligible else "X"
        ax_all.scatter(
            x, float(row[metric.column]), s=52 if eligible else 70,
            c=color if eligible else "#AAAAAA", marker=marker,
            edgecolors="black", linewidths=0.5, zorder=3,
        )
        if fid in selected_good:
            ax_all.scatter(
                x, float(row[metric.column]), s=150, facecolors="none",
                edgecolors=DISPLAY_PURE_COLOR, linewidths=2.2, zorder=4,
            )
        if fid in selected_bad:
            ax_all.scatter(
                x, float(row[metric.column]), s=205, facecolors="none",
                edgecolors=DISPLAY_NOMINAL_COLOR, linewidths=2.2,
                marker="s", zorder=4,
            )
    ax_all.set_xticks(np.arange(len(finite)))
    ax_all.set_xticklabels(
        [short_id(fid) for fid in finite["flight_id"]], rotation=55,
        ha="right", fontsize=8,
    )
    ax_all.set_title("All 21 recorded sessions\n(X = censored/incomplete; rings = selected)")
    ax_all.set_ylabel(f"{metric.title} [{metric.unit}]")
    ax_all.grid(axis="y", alpha=0.22)
    handles = [
        Line2D([0], [0], marker="o", linestyle="", color="none",
               markerfacecolor=color, markeredgecolor="black", label=condition)
        for condition, color in CONDITION_COLORS.items()
    ]
    ax_all.legend(handles=handles, loc="best", frameon=False, fontsize=8)

    pair_rows = list(selection["pairs"])
    pure_values = np.asarray([
        float(pair["display_pure_value"]) for pair in pair_rows
    ])
    nominal_values = np.asarray([
        float(pair["display_nominal_value"]) for pair in pair_rows
    ])
    all_pair_values = np.r_[pure_values, nominal_values]
    value_range = max(float(np.ptp(all_pair_values)),
                      abs(float(np.mean(all_pair_values))) * 0.12, 1e-6)
    y_lower = float(np.min(all_pair_values) - 0.13 * value_range)
    y_upper = float(np.max(all_pair_values) + 0.13 * value_range)
    min_label_gap = 0.075 * (y_upper - y_lower)
    pure_label_y = spread_label_positions(
        pure_values, y_lower, y_upper, min_label_gap
    )
    nominal_label_y = spread_label_positions(
        nominal_values, y_lower, y_upper, min_label_gap
    )
    for pair_index, pair in enumerate(pair_rows):
        y0 = float(pair["display_pure_value"])
        y1 = float(pair["display_nominal_value"])
        ax_pairs.plot([0, 1], [y0, y1], color="#999999", linewidth=1.3, alpha=0.8)
        ax_pairs.scatter([0], [y0], s=70, c=DISPLAY_PURE_COLOR,
                         edgecolors="black", linewidths=0.5, zorder=3)
        ax_pairs.scatter([1], [y1], s=70, c=DISPLAY_NOMINAL_COLOR,
                         edgecolors="black", linewidths=0.5, zorder=3)
        left_label = (
            f"{pair['display_pure_short_id']}"
            f" ({pair['display_pure_recorded_condition']})"
        )
        right_label = (
            f"{pair['display_nominal_short_id']}"
            f" ({pair['display_nominal_recorded_condition']})"
        )
        ax_pairs.annotate(
            left_label, xy=(0, y0), xytext=(-0.09, pure_label_y[pair_index]),
            textcoords="data", ha="right", va="center", fontsize=7.5,
            arrowprops={"arrowstyle": "-", "color": "#777777", "lw": 0.55},
        )
        ax_pairs.annotate(
            right_label, xy=(1, y1), xytext=(1.09, nominal_label_y[pair_index]),
            textcoords="data", ha="left", va="center", fontsize=7.5,
            arrowprops={"arrowstyle": "-", "color": "#777777", "lw": 0.55},
        )
    pure_mean = float(selection["display_pure_mean"])
    nominal_mean = float(selection["display_nominal_mean"])
    ax_pairs.scatter([0, 1], [pure_mean, nominal_mean], s=190, marker="D",
                     c=[DISPLAY_PURE_COLOR, DISPLAY_NOMINAL_COLOR],
                     edgecolors="black", linewidths=1.0, zorder=5)
    ax_pairs.set_xlim(-0.72, 1.72)
    ax_pairs.set_ylim(y_lower, y_upper)
    ax_pairs.set_xticks([0, 1])
    ax_pairs.set_xticklabels(["display-PURE\ncandidates", "display-Nominal\ncandidates"])
    ax_pairs.set_title(
        "Five matched, disjoint pairs\n"
        f"means {metric_formatter(metric, pure_mean)} -> "
        f"{metric_formatter(metric, nominal_mean)}; "
        f"gap={selection['favorable_separation_pool_sd']:.2f} pool SD"
    )
    ax_pairs.grid(axis="y", alpha=0.22)
    fig.suptitle(
        f"Metric-optimized visualization candidates: {metric.title}\n"
        "Recorded conditions are shown in parentheses and were not used for selection",
        fontsize=13,
    )
    fig.text(
        0.5, 0.012,
        "Matching: path ratio <=1.20, duration ratio <=1.25, position-opportunity "
        f"gap <=0.10; arm balance <= {selection['arm_balance_delta_smd']:.2f} SD",
        ha="center", fontsize=8.5, color="#444444",
    )
    fig.tight_layout(rect=(0, 0.045, 1, 0.92))
    stem = metric.column
    fig.savefig(plots_dir / f"{stem}.png", dpi=220, transparent=False)
    fig.savefig(plots_dir / f"{stem}.svg", transparent=True)
    plt.close(fig)


def plot_membership_heatmap(
    master: pd.DataFrame,
    selections: Sequence[Mapping[str, Any]],
    output: Path,
) -> None:
    ids = master["flight_id"].tolist()
    matrix = np.zeros((len(ids), len(selections)), dtype=float)
    id_index = {flight_id: index for index, flight_id in enumerate(ids)}
    for column, selection in enumerate(selections):
        for pair in selection["pairs"]:
            matrix[id_index[pair["display_pure_flight_id"]], column] = -1.0
            matrix[id_index[pair["display_nominal_flight_id"]], column] = 1.0
    fig, ax = plt.subplots(figsize=(16.0, 8.2))
    from matplotlib.colors import ListedColormap
    cmap = ListedColormap([DISPLAY_PURE_COLOR, "#F7F7F7", DISPLAY_NOMINAL_COLOR])
    ax.imshow(matrix, aspect="auto", cmap=cmap, vmin=-1, vmax=1)
    ax.set_yticks(np.arange(len(ids)))
    ax.set_yticklabels([
        f"{short_id(fid)} ({master.loc[master.flight_id == fid, 'recorded_condition'].iloc[0]})"
        for fid in ids
    ], fontsize=8)
    ax.set_xticks(np.arange(len(selections)))
    ax.set_xticklabels(
        [selection["metric"]["title"] for selection in selections],
        rotation=50, ha="right", fontsize=8,
    )
    ax.set_title(
        "Metric-specific candidate membership\n"
        "purple=display-PURE, gold=display-Nominal, white=not selected"
    )
    ax.set_xticks(np.arange(-0.5, len(selections), 1), minor=True)
    ax.set_yticks(np.arange(-0.5, len(ids), 1), minor=True)
    ax.grid(which="minor", color="white", linewidth=0.5)
    ax.tick_params(which="minor", bottom=False, left=False)
    fig.tight_layout()
    fig.savefig(output.with_suffix(".png"), dpi=220)
    fig.savefig(output.with_suffix(".svg"), transparent=True)
    plt.close(fig)


def write_excel(
    output: Path,
    master: pd.DataFrame,
    pair_frame: pd.DataFrame,
    summary_frame: pd.DataFrame,
    definitions: pd.DataFrame,
) -> None:
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        master.to_excel(writer, sheet_name="all21_master", index=False)
        summary_frame.to_excel(writer, sheet_name="candidate_summary", index=False)
        pair_frame.to_excel(writer, sheet_name="candidate_pairs", index=False)
        definitions.to_excel(writer, sheet_name="metric_definitions", index=False)
        workbook = writer.book
        for sheet in workbook.worksheets:
            sheet.freeze_panes = "A2"
            sheet.auto_filter.ref = sheet.dimensions
            fill = PatternFill("solid", fgColor="D9EAF7")
            for cell in sheet[1]:
                cell.font = Font(bold=True)
                cell.fill = fill
                cell.alignment = Alignment(wrap_text=True, vertical="top")
            for column_index, cells in enumerate(sheet.columns, start=1):
                sample = [str(cell.value) if cell.value is not None else "" for cell in list(cells)[:40]]
                width = min(55, max(10, max(map(len, sample), default=10) + 2))
                sheet.column_dimensions[get_column_letter(column_index)].width = width


def write_index_html(
    output: Path,
    master: pd.DataFrame,
    selections: Sequence[Mapping[str, Any]],
    summary_frame: pd.DataFrame,
) -> None:
    cards = []
    for selection in selections:
        metric = selection["metric"]
        stem = metric["column"]
        pairs = pd.DataFrame(selection["pairs"])[[
            "display_pure_short_id", "display_pure_recorded_condition",
            "display_pure_value", "display_nominal_short_id",
            "display_nominal_recorded_condition", "display_nominal_value",
        ]]
        cards.append(f"""
        <section>
          <h2>{html.escape(metric['title'])}</h2>
          <a href="plots/{stem}.svg"><img src="plots/{stem}.svg" alt="{stem}"></a>
          {pairs.to_html(index=False, border=0, classes='pairs', float_format=lambda x: f'{x:.5g}')}
        </section>
        """)
    document = f"""<!doctype html>
<html lang="en"><head><meta charset="utf-8"><title>21-flight metric candidates</title>
<style>
body{{font-family:Arial,sans-serif;margin:24px;color:#222;max-width:1500px}}
h1,h2{{margin-bottom:.35em}} .warning{{background:#fff3cd;border:1px solid #e5c15a;padding:12px}}
img{{width:100%;max-width:1400px;border:1px solid #ddd}} section{{margin:42px 0}}
table{{border-collapse:collapse;font-size:13px}} th,td{{border:1px solid #ddd;padding:5px 8px}}
th{{background:#eaf2f8;position:sticky;top:0}} code{{background:#eee;padding:2px 4px}}
</style></head><body>
<h1>Metric-optimized 5-vs-5 visualization candidates</h1>
<p class="warning"><b>Interpretation:</b> display-PURE/display-Nominal are metric-optimized
presentation roles. They do not replace the immutable recorded condition shown beside every ID.
Near-crash or incomplete flights remain in the master table but are excluded from selection.</p>
<p>Common window: stable hover through landing. Estimator: one frozen full-LIVO hybrid-IMU
configuration for all 21 sessions. Offline cloth endpoint is geometric view fraction <i>f</i>, not
recorded online S.</p>
<p><a href="all21_final_metrics.csv">Master CSV</a> | <a href="final_metric_table_and_candidates.xlsx">Excel workbook</a> |
<a href="candidate_sets.json">Machine-readable manifest</a> | <a href="plots/selection_membership.svg">Membership heatmap</a></p>
<h2>Candidate summary</h2>
{summary_frame.to_html(index=False, border=0, float_format=lambda x: f'{x:.5g}')}
{''.join(cards)}
</body></html>"""
    output.write_text(document, encoding="utf-8")


def write_readme(
    output: Path,
    metadata: Mapping[str, Any],
    summary: pd.DataFrame,
) -> None:
    compact_columns = [
        "metric", "display_pure_mean", "display_nominal_mean",
        "favorable_separation_pool_sd", "balance_delta_smd",
    ]
    compact = summary[compact_columns]
    markdown_rows = [
        "| " + " | ".join(compact_columns) + " |",
        "| " + " | ".join(["---"] * len(compact_columns)) + " |",
    ]
    for _, row in compact.iterrows():
        values = []
        for column in compact_columns:
            value = row[column]
            values.append(f"{float(value):.4g}" if isinstance(
                value, (float, np.floating)
            ) else str(value))
        markdown_rows.append("| " + " | ".join(values) + " |")
    lines = [
        "# 21-flight metric candidate pack",
        "",
        "This directory is regenerated by `tools/fastlivo/build_metric_optimized_candidate_sets.py`.",
        "",
        "## What the labels mean",
        "",
        "`display-PURE` and `display-Nominal` are metric-optimized visualization roles. ",
        "The immutable recorded condition is retained in every table/plot and is not used by the optimizer.",
        "",
        f"- Master rows: {metadata['session_count']}",
        f"- Default eligible completed rows: {metadata['eligible_primary_count']}",
        f"- Censored/incomplete but retained: {', '.join(metadata['incomplete_flight_ids'])}",
        "- Common window: stable hover to landing",
        "- Estimator: full_livo_hybrid_imu_acc10_hover_r1, no spatial alignment",
        "- Cloth endpoint: offline fixed-AABB/FOV view fraction f, not online S",
        "",
        "## Matching and optimization",
        "",
        "Each metric independently selects five vertex-disjoint directed pairs. Hard pair calipers are ",
        "GT path ratio <=1.20, duration ratio <=1.25, and position-only cloth opportunity gap <=0.10. ",
        "Arm-level standardized differences in log path, log duration, and opportunity are <=0.10 SD ",
        "when feasible (the manifest records any deterministic relaxation). The objective maximizes the ",
        "favorable standardized metric gap.",
        "",
        "## Files",
        "",
        "- `all21_final_metrics.csv/json`: wide master table",
        "- `final_metric_table_and_candidates.xlsx`: Windows-friendly workbook",
        "- `candidate_sets.json`, `candidate_pairs.csv`, `candidate_summary.csv`: selections",
        "- `plots/*.svg` and `plots/*.png`: one figure per metric plus membership heatmap",
        "- `index.html`: offline browser index",
        "- `provenance.json`, `metric_definitions.csv`: formulas and input hashes",
        "",
        "## Compact summary",
        "",
        "\n".join(markdown_rows),
        "",
    ]
    output.write_text("\n".join(lines), encoding="utf-8")


def main() -> None:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--output", type=Path, default=DEFAULT_OUTPUT)
    parser.add_argument(
        "--opportunity-yaw-step-deg", type=float, default=5.0,
        help="yaw grid for position-only cloth opportunity (default: 5 deg)",
    )
    args = parser.parse_args()
    output = args.output.resolve()
    output.mkdir(parents=True, exist_ok=True)
    plots_dir = output / "plots"
    plots_dir.mkdir(parents=True, exist_ok=True)

    master, metadata = build_master(output, args.opportunity_yaw_step_deg)
    master.to_csv(output / "all21_final_metrics.csv", index=False)
    atomic_json(output / "all21_final_metrics.json", {
        "schema": "campaign21_final_metrics/v1",
        "metadata": metadata,
        "rows": json_records(master),
    })

    selections: List[Dict[str, Any]] = []
    all_pairs: List[Dict[str, Any]] = []
    summary_rows: List[Dict[str, Any]] = []
    for index, metric in enumerate(METRICS, start=1):
        print(f"[select {index:02d}/{len(METRICS)}] {metric.column}", flush=True)
        selection = optimize_pairs(master, metric)
        selections.append(selection)
        all_pairs.extend(selection["pairs"])
        summary_rows.append({
            "metric": metric.column,
            "title": metric.title,
            "unit": metric.unit,
            "higher_is_better": metric.higher_is_better,
            "display_pure_mean": selection["display_pure_mean"],
            "display_nominal_mean": selection["display_nominal_mean"],
            "favorable_separation_pool_sd": selection[
                "favorable_separation_pool_sd"
            ],
            "balance_delta_smd": selection["arm_balance_delta_smd"],
            "path_smd": selection["covariate_smd"]["log_gt_path"],
            "duration_smd": selection["covariate_smd"]["log_duration"],
            "opportunity_smd": selection["covariate_smd"]["position_opportunity"],
        })
        plot_metric(master, metric, selection, plots_dir)
    pair_frame = pd.DataFrame(all_pairs)
    summary_frame = pd.DataFrame(summary_rows)
    pair_frame.to_csv(output / "candidate_pairs.csv", index=False)
    summary_frame.to_csv(output / "candidate_summary.csv", index=False)
    atomic_json(output / "candidate_sets.json", {
        "schema": "metric_optimized_candidate_sets/v1",
        "selection_role_warning": (
            "display-PURE/display-Nominal are exploratory metric-optimized roles; "
            "recorded_condition is immutable and was not used for selection"
        ),
        "common_candidate_pool": master.loc[
            master["eligible_primary_selection"], "flight_id"
        ].tolist(),
        "selections": selections,
    })
    definitions = pd.DataFrame([asdict(metric) for metric in METRICS])
    definitions.to_csv(output / "metric_definitions.csv", index=False)
    write_excel(
        output / "final_metric_table_and_candidates.xlsx",
        master, pair_frame, summary_frame, definitions,
    )
    plot_membership_heatmap(master, selections, plots_dir / "selection_membership")
    write_index_html(output / "index.html", master, selections, summary_frame)

    input_paths = [
        SPEC_PATH, RESULTS_PATH, FUSION_PATH, CLOTH_PATH, SCENARIO_PATH,
        CACHE_INDEX_PATH, INPUT_QC_PATH, Path(__file__).resolve(),
        HERE / "plot_flight_timeseries.py", HERE / "eval_fastlivo.py",
    ]
    provenance = {
        "schema": "campaign21_metric_candidate_provenance/v1",
        "generator": str(Path(__file__).resolve()),
        "generator_sha256": sha256(Path(__file__).resolve()),
        "common_window": "stable-hover to landing",
        "spatial_alignment": "none",
        "association_max_diff_s": MAX_ASSOCIATION_DIFF_S,
        "opportunity": {
            "definition": "max offline cloth-view fraction over yaw grid at fixed GT position",
            "yaw_step_deg": args.opportunity_yaw_step_deg,
            "position_hz": 10.0,
        },
        "eligibility": {
            "default": (
                "valid estimator + feature span >=0.95 + cloth available + "
                "start/TERMINAL_NAV/DONE/landing + terminal GT within 0.2m"
            ),
            **metadata,
        },
        "inputs": [
            {"path": str(path), "size_bytes": path.stat().st_size,
             "sha256": sha256(path)}
            for path in input_paths
        ],
        "metric_definitions": [asdict(metric) for metric in METRICS],
    }
    atomic_json(output / "provenance.json", provenance)
    write_readme(output / "README.md", metadata, summary_frame)

    checksum_files = sorted(
        path for path in output.rglob("*")
        if path.is_file() and path.name != "sha256sums.txt"
    )
    with (output / "sha256sums.txt").open("w", encoding="utf-8") as stream:
        for path in checksum_files:
            stream.write(f"{sha256(path)}  {path.relative_to(output)}\n")
    print(f"[done] {output}")
    print(f"[done] master={len(master)} rows, eligible={metadata['eligible_primary_count']}")
    print(f"[done] metrics={len(METRICS)}, pairs={len(pair_frame)}")


if __name__ == "__main__":
    main()
